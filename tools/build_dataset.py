#!/usr/bin/env python3
"""Weekly Marketwatch dataset builder.

Turns a Home Builders Research "WTSW" Excel drop into the app's SUBS data,
carrying over enrichment (address/lat-lng/website/mapUrl) from the currently
deployed data so only NEW communities need research. See tools/DATASET_UPDATE.md.

Two-step flow:

  1) python tools/build_dataset.py parse "WTSW 260802.xlsx"
       - parses the workbook, maps columns/areas, canonicalizes builder names
       - carries over enrichment + STABLE ids from the current index.html
       - writes tools/_work/dataset.json (full set; new communities blank)
       - writes per-builder enrichment inputs to tools/_work/enrich_in/*.json
       - prints the worklist of communities still needing enrichment

     -> run the research agents on tools/_work/enrich_in/*.json, having each
        write tools/_work/enrich_out/<batch>.json (same schema as enrich_in
        plus address/website/webDomain/webType/lat/lng/geoNote/confidence).

  2) python tools/build_dataset.py apply
       - merges tools/_work/enrich_out/*.json into the dataset
       - verifies every coordinate (Southern NV bounds + near its area center)
       - injects SUBS into index.html and updates BUILD_DATE / footer week-ending
       - prints a review list of medium/low-confidence or flagged rows

Then: python tools/verify.py  (must be green) -> commit + push.

Requires: openpyxl  (pip install openpyxl)
"""
import sys, os, re, json, glob, math, datetime

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
WORK = os.path.join(REPO, "tools", "_work")
ENRICH_IN = os.path.join(WORK, "enrich_in")
ENRICH_OUT = os.path.join(WORK, "enrich_out")
INDEX = os.path.join(REPO, "index.html")

AREA = {"NW":"Northwest","H":"Henderson","SW":"Southwest","NLV":"North Las Vegas",
        "SO":"South","E":"East","BC":"Boulder City","MSQ":"Mesquite","P":"Pahrump"}
AREA_CENTER = {"Northwest":(36.25,-115.28),"Henderson":(36.0,-114.98),"Southwest":(36.0,-115.28),
               "North Las Vegas":(36.25,-115.12),"South":(36.0,-115.15),"East":(36.13,-115.05),
               "Boulder City":(35.97,-114.83),"Mesquite":(36.80,-114.07),"Pahrump":(36.21,-115.98)}
BUILDER_CANON = {"BEAZER HOMES":"Beazer Homes","CENTURY COMMUNITIES":"Century Communities",
    "CONTOUR HOMES":"Contour Homes","DR HORTON":"D.R. Horton","KB HOME":"KB Home","LENNAR":"Lennar",
    "PINNACLE HOMES":"Pinnacle Homes","PULTE GROUP":"Pulte Group","SEKISUI HOUSE US":"Sekisui House",
    "SHAWOOD":"Shawood","SHEA HOMES":"Shea Homes","SIGNATURE HOMES":"Signature Homes",
    "SUMMIT HOMES":"Summit Homes","TAYLOR MORRISON":"Taylor Morrison","TOLL BROTHERS":"Toll Brothers",
    "TOUCHSTONE LIVING":"Touchstone Living","TRI POINTE HOMES":"Tri Pointe Homes"}
# app field  ->  raw WTSW column index (0-based)
COL = dict(builder=0, sub=1, area=2, ptype=3, traffic=4, ytd=9, totalLot=10,
           unsold=12, opened=14, standing=18, minSqft=19, maxSqft=20, lo=21, hi=22, mp=23)
KEYS = ["id","builder","sub","area","mp","type","lo","hi","minSqft","maxSqft","standing","unsold",
        "ytd","traffic","totalLot","opened","address","lat","lng","mapUrl","website","webDomain",
        "webType","geoNote"]

def norm(x): return re.sub(r"[^a-z0-9]", "", str(x).lower())
def tc(s): return " ".join(w.capitalize() for w in str(s).split())
def canon_builder(b): return BUILDER_CANON.get(str(b).strip().upper(), tc(b))
def clean_sub(s):
    for a,b in [("(wh)","(WH)"),("(rah)","(RAH)"),("(dw)","(DW)"),("(aw)","(AW)"),("(sbh)","(SBH)")]:
        s=s.replace(a,b)
    s=re.sub(r"\bIii\b","III",s); s=re.sub(r"\bIi\b","II",s); s=re.sub(r"\bBy Tm\b","by TM",s)
    return s
def price(v):
    if v is None: return None
    d=re.sub(r"[^0-9]","",str(v)); return int(d) if d else None
def num(v):
    if isinstance(v,float) and v==int(v): return int(v)
    return v
def hav(a,b,c,d):
    R=6371;p=math.pi/180
    x=math.sin((c-a)*p/2)**2+math.cos(a*p)*math.cos(c*p)*math.sin((d-b)*p/2)**2
    return 2*R*math.asin(min(1,math.sqrt(x)))

def load_current_subs():
    html=open(INDEX,encoding="utf-8").read()
    m=re.search(r"const SUBS = (\[.*?\}\]);", html, re.S)
    return json.loads(m.group(1)) if m else []

def read_wtsw(path):
    import openpyxl
    wb=openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws=None; datestr=None
    for name in wb.sheetnames:
        if name.upper().startswith("WTSW"):
            ws=wb[name]
            mm=re.search(r"(\d{8})", name)
            if mm: datestr=mm.group(1)
            break
    if ws is None: sys.exit("No WTSW sheet found.")
    rows=[r for r in ws.iter_rows(values_only=True)][1:]
    recs=[]
    for r in rows:
        if not (r[COL['builder']] and r[COL['sub']]): continue
        if norm(r[COL['sub']])=="subtotals": continue
        recs.append(r)
    return recs, datestr

def parse_mode(xlsx):
    recs, datestr = read_wtsw(xlsx)
    cur=load_current_subs()
    cur_bn={(norm(s['builder']),norm(s['sub'])):s for s in cur}
    cur_n={norm(s['sub']):s for s in cur}
    seen=set(); out=[]; need=[]
    for r in recs:
        builder=canon_builder(r[COL['builder']]); sub=clean_sub(tc(r[COL['sub']]))
        code=str(r[COL['area']]).strip() if r[COL['area']] else ""
        ex=cur_bn.get((norm(builder),norm(sub))) or cur_n.get(norm(sub))
        if ex: cid=ex['id']
        else:
            base=re.sub(r"[^a-z0-9]+","-",sub.lower()).strip("-"); cid=base; i=2
            while cid in seen: cid=f"{base}-{i}"; i+=1
        seen.add(cid)
        rec={"id":cid,"builder":builder,"sub":sub,"area":AREA.get(code,code or "Unknown"),
             "mp":clean_sub(tc(r[COL['mp']])) if r[COL['mp']] else "",
             "type":"Attached" if str(r[COL['ptype']]).strip()=="ATT" else ("Detached" if str(r[COL['ptype']]).strip()=="DET" else ""),
             "lo":price(r[COL['lo']]),"hi":price(r[COL['hi']]),
             "minSqft":num(r[COL['minSqft']]),"maxSqft":num(r[COL['maxSqft']]),
             "standing":num(r[COL['standing']]),"unsold":num(r[COL['unsold']]),"ytd":num(r[COL['ytd']]),
             "traffic":num(r[COL['traffic']]),"totalLot":num(r[COL['totalLot']]),
             "opened":str(r[COL['opened']]) if r[COL['opened']] else "",
             "address": ex.get('address','') if ex else "", "lat": ex.get('lat') if ex else None,
             "lng": ex.get('lng') if ex else None, "mapUrl": ex.get('mapUrl','') if ex else "",
             "website": ex.get('website','') if ex else "", "webDomain": ex.get('webDomain','') if ex else "",
             "webType": ex.get('webType','') if ex else "", "geoNote": ex.get('geoNote','') if ex else ""}
        if not (ex and ex.get('lat') is not None):
            need.append(rec)
        out.append(rec)
    os.makedirs(ENRICH_IN, exist_ok=True); os.makedirs(ENRICH_OUT, exist_ok=True)
    for f in glob.glob(os.path.join(ENRICH_IN,"*.json")): os.remove(f)
    json.dump({"datestr":datestr,"records":out}, open(os.path.join(WORK,"dataset.json"),"w",encoding="utf-8"), ensure_ascii=False, indent=1)
    # per-builder enrichment inputs (split >12 into batches of 10)
    byb={}
    for x in need: byb.setdefault(x['builder'],[]).append(x)
    nb=0
    for builder,items in byb.items():
        chunks=[items[i:i+10] for i in range(0,len(items),10)] if len(items)>12 else [items]
        for j,part in enumerate(chunks,1):
            bid=re.sub(r"[^a-z0-9]+","-",builder.lower()).strip("-")+(f"-{j}" if len(chunks)>1 else "")
            json.dump([{"sub":x['sub'],"builder":builder,"area":x['area'],"type":x['type'],"lo":x['lo'],"hi":x['hi'],"mp":x['mp']} for x in part],
                      open(os.path.join(ENRICH_IN,f"in_{bid}.json"),"w",encoding="utf-8"), ensure_ascii=False, indent=1)
            nb+=1
    print(f"Parsed {len(out)} communities (week {datestr}).  carried-over={len(out)-len(need)}  need-enrichment={len(need)}")
    print(f"Wrote {WORK}/dataset.json")
    if need:
        print(f"\n{nb} enrichment batch input(s) in {ENRICH_IN}/ . Run research agents; write results to {ENRICH_OUT}/out_<batch>.json, then: python tools/build_dataset.py apply")
        import collections
        for b,c in collections.Counter(x['builder'] for x in need).most_common(): print(f"   {c:>3}  {b}")
    else:
        print("\nNothing needs enrichment (all carried over). Run: python tools/build_dataset.py apply")

def apply_mode():
    ds=json.load(open(os.path.join(WORK,"dataset.json"),encoding="utf-8"))
    recs=ds["records"]; datestr=ds.get("datestr")
    enr={}
    for f in glob.glob(os.path.join(ENRICH_OUT,"*.json")):
        try:
            for e in json.load(open(f,encoding="utf-8")):
                if e.get("sub"): enr[norm(e["sub"])]=e
        except Exception as ex: print("skip",f,ex)
    review=[]
    for s in recs:
        if s.get("lat") is None:
            e=enr.get(norm(s["sub"]))
            if e:
                for k in ["address","website","webDomain","webType","geoNote"]: s[k]=e.get(k,"") or ""
                s["lat"]=e.get("lat"); s["lng"]=e.get("lng"); s["_conf"]=e.get("confidence","?")
            else: review.append((s["sub"],s["builder"],"NO enrichment")); continue
        if isinstance(s.get("lat"),(int,float)) and isinstance(s.get("lng"),(int,float)):
            s["mapUrl"]=f"https://www.google.com/maps/search/?api=1&query={s['lat']},{s['lng']}"
            lat,lng=s["lat"],s["lng"]
            if not (35.0<=lat<=37.2 and -116.6<=lng<=-113.9): review.append((s["sub"],s["builder"],f"OUT OF BOUNDS {lat:.3f},{lng:.3f}"))
            else:
                c=AREA_CENTER.get(s["area"])
                if c and hav(lat,lng,c[0],c[1])>40: review.append((s["sub"],s["builder"],f"{hav(lat,lng,c[0],c[1]):.0f}km from {s['area']} center"))
            if s.get("_conf") in ("medium","low"): review.append((s["sub"],s["builder"],f"confidence={s['_conf']}"))
        else:
            review.append((s["sub"],s["builder"],"NO COORDS"))
    clean=[{k:s.get(k) for k in KEYS} for s in recs]
    inline="const SUBS = "+json.dumps(clean,ensure_ascii=False,separators=(", ",": "))+";"
    html=open(INDEX,encoding="utf-8").read()
    html=re.sub(r"const SUBS = \[.*?\}\];", inline.replace("\\","\\\\"), html, count=1, flags=re.S)
    if datestr:
        d=datetime.date(int(datestr[:4]),int(datestr[4:6]),int(datestr[6:8]))
        html=re.sub(r'const BUILD_DATE = "[^"]*";', f'const BUILD_DATE = "{d.strftime("%b %-d, %Y") if os.name!="nt" else d.strftime("%b %#d, %Y")}";', html, count=1)
        html=re.sub(r"Data week ending \d\d/\d\d/\d{4}", f"Data week ending {d.strftime('%m/%d/%Y')}", html, count=1)
    open(INDEX,"w",encoding="utf-8",newline="").write(html)
    print(f"Injected {len(clean)} communities into index.html (week {datestr}).")
    print(f"\n=== REVIEW LIST ({len(review)}) ===")
    for sub,b,why in review: print(f"  {sub:<30} {b:<20} {why}")
    print("\nNext: python tools/verify.py  (must be green) -> commit + push.")

if __name__=="__main__":
    if len(sys.argv)>=3 and sys.argv[1]=="parse": parse_mode(sys.argv[2])
    elif len(sys.argv)>=2 and sys.argv[1]=="apply": apply_mode()
    else: print(__doc__)
